"""
文件解析服务
支持: PDF, DOCX, XLSX, XLS, TXT, MD, CSV, JSON, PNG, JPG, JPEG
"""
import io
import logging
from typing import Tuple, Dict, Any, Optional
from pathlib import Path

from app.config import settings

logger = logging.getLogger(__name__)


class FileParseError(Exception):
    """文件解析错误"""
    pass


class FileParser:
    """统一文件解析服务"""

    def __init__(self):
        self.max_file_size = settings.MAX_FILE_SIZE
        self.allowed_extensions = settings.ALLOWED_EXTENSIONS

    def get_supported_types(self) -> set:
        """获取支持的文件类型"""
        return self.allowed_extensions

    @property
    def max_file_size_mb(self) -> int:
        """获取最大文件大小（MB）"""
        return self.max_file_size // (1024 * 1024)

    async def parse(
        self,
        content: bytes,
        file_type: str,
        filename: str = "unknown"
    ) -> Tuple[str, Dict[str, Any]]:
        """
        解析文件内容

        Args:
            content: 文件二进制内容
            file_type: 文件扩展名 (不含点)
            filename: 原始文件名

        Returns:
            (提取的文本, 元数据字典)
        """
        # 检查文件类型
        file_type = file_type.lower().lstrip('.')
        if file_type not in self.allowed_extensions:
            raise FileParseError(f"不支持的文件格式: .{file_type}")

        # 检查文件大小
        if len(content) > self.max_file_size:
            raise FileParseError(f"文件大小超过 {self.max_file_size // (1024*1024)}MB 限制")

        # 根据类型分发处理
        parsers = {
            'txt': self._parse_text,
            'md': self._parse_text,
            'csv': self._parse_text,
            'json': self._parse_text,
            'pdf': self._parse_pdf,
            'docx': self._parse_docx,
            'xlsx': self._parse_excel,
            'xls': self._parse_excel,
            'png': self._parse_image,
            'jpg': self._parse_image,
            'jpeg': self._parse_image,
        }

        parser = parsers.get(file_type)
        if not parser:
            raise FileParseError(f"未实现 {file_type} 格式的解析器")

        try:
            text, metadata = await parser(content, file_type)
            metadata['fileName'] = filename
            metadata['fileSize'] = len(content)
            metadata['fileType'] = file_type

            logger.info(f"文件解析成功: {filename}, 类型: {file_type}, 文本长度: {len(text)}")
            return text, metadata

        except FileParseError:
            raise
        except Exception as e:
            logger.error(f"文件解析失败: {filename}, 错误: {str(e)}")
            raise FileParseError(f"文件解析失败: {str(e)}")

    async def _parse_text(self, content: bytes, file_type: str) -> Tuple[str, Dict[str, Any]]:
        """解析纯文本文件"""
        try:
            # 尝试 UTF-8 解码
            text = content.decode('utf-8')
        except UnicodeDecodeError:
            try:
                # 尝试 GBK 解码 (中文 Windows)
                text = content.decode('gbk')
            except UnicodeDecodeError:
                # 尝试 Latin-1
                text = content.decode('latin-1')

        return text.strip(), {'isOCR': False}

    async def _parse_pdf(self, content: bytes, file_type: str) -> Tuple[str, Dict[str, Any]]:
        """解析 PDF 文件"""
        try:
            import pdfplumber
        except ImportError:
            raise FileParseError("PDF 解析库未安装，请运行: pip install pdfplumber")

        try:
            text_parts = []
            page_count = 0

            with pdfplumber.open(io.BytesIO(content)) as pdf:
                page_count = len(pdf.pages)

                for i, page in enumerate(pdf.pages):
                    page_text = page.extract_text() or ""
                    if page_text.strip():
                        text_parts.append(f"[第{i+1}页]\n{page_text}")

            text = "\n\n".join(text_parts).strip()

            if len(text) < 50:
                raise FileParseError(
                    "此 PDF 文字内容过少，可能是扫描版。"
                    "建议：1) 复制文字直接粘贴；2) 或截图后以图片格式上传"
                )

            return text, {'pageCount': page_count, 'isOCR': False}

        except FileParseError:
            raise
        except Exception as e:
            raise FileParseError(f"PDF 解析失败: {str(e)}")

    async def _parse_docx(self, content: bytes, file_type: str) -> Tuple[str, Dict[str, Any]]:
        """解析 DOCX 文件"""
        try:
            import mammoth
        except ImportError:
            raise FileParseError("DOCX 解析库未安装，请运行: pip install mammoth")

        try:
            result = mammoth.extract_raw_text(io.BytesIO(content))
            text = result.value.strip()

            if not text:
                raise FileParseError("Word 文档内容为空")

            # 估算页数
            page_count = max(1, len(text) // 500)

            return text, {'pageCount': page_count, 'isOCR': False}

        except FileParseError:
            raise
        except Exception as e:
            raise FileParseError(f"Word 文档解析失败，请确保是 .docx 格式: {str(e)}")

    async def _parse_excel(self, content: bytes, file_type: str) -> Tuple[str, Dict[str, Any]]:
        """解析 Excel 文件"""
        try:
            import openpyxl
            import xlrd
        except ImportError as e:
            raise FileParseError(f"Excel 解析库未安装: {str(e)}")

        try:
            texts = []
            sheet_count = 0

            if file_type == 'xlsx':
                # 使用 openpyxl 处理 .xlsx
                wb = openpyxl.load_workbook(io.BytesIO(content))
                sheet_count = len(wb.sheetnames)

                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    texts.append(f"\n【工作表: {sheet_name}】\n")

                    for row in sheet.iter_rows(values_only=True):
                        row_values = [str(cell) if cell is not None else '' for cell in row]
                        if any(v.strip() for v in row_values):
                            texts.append(' | '.join(row_values))

            else:
                # 使用 xlrd 处理 .xls
                wb = xlrd.open_workbook(file_contents=content)
                sheet_count = wb.nsheets

                for sheet_idx in range(wb.nsheets):
                    sheet = wb.sheet_by_index(sheet_idx)
                    texts.append(f"\n【工作表: {sheet.name}】\n")

                    for row_idx in range(sheet.nrows):
                        row_values = [str(sheet.cell_value(row_idx, col_idx)) for col_idx in range(sheet.ncols)]
                        if any(v.strip() for v in row_values):
                            texts.append(' | '.join(row_values))

            text = '\n'.join(texts).strip()

            if not text:
                raise FileParseError("Excel 文件内容为空")

            return text, {'sheetCount': sheet_count, 'isOCR': False}

        except FileParseError:
            raise
        except Exception as e:
            raise FileParseError(f"Excel 文件解析失败: {str(e)}")

    async def _parse_image(self, content: bytes, file_type: str) -> Tuple[str, Dict[str, Any]]:
        """OCR 图片识别"""
        try:
            import pytesseract
            from PIL import Image
        except ImportError as e:
            raise FileParseError(f"OCR 库未安装: {str(e)}")

        try:
            image = Image.open(io.BytesIO(content))

            # 使用中英文识别
            text = pytesseract.image_to_string(image, lang='chi_sim+eng')
            text = text.strip()

            if not text:
                raise FileParseError("未能从图片中识别出文字，请尝试：1) 复制文字直接粘贴；2) 或使用更清晰的图片")

            return text, {'isOCR': True, 'imageSize': image.size}

        except FileParseError:
            raise
        except Exception as e:
            raise FileParseError(f"图片识别失败: {str(e)}")


# 单例实例
file_parser = FileParser()
