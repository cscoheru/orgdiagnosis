"""
Excel 报告生成器 — 模板数据填充

刚性约束:
1. 打开现有模板文件，不破坏原有格式和公式
2. 从指定行开始写入数据
3. 保留原有单元格样式
"""
import os
from typing import Any
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet


def generate_compensation_excel(
    template_path: str,
    data: list[dict[str, Any]],
    output_path: str,
    start_row: int = 3,
    column_mapping: dict[str, str] | None = None,
) -> dict[str, Any]:
    """生成薪酬 Excel 报告"""
    if column_mapping is None:
        column_mapping = {
            "position_name": "A",
            "salary_level": "B",
            "base_salary": "C",
            "bonus": "D",
            "total_compensation": "E",
        }

    try:
        if not os.path.exists(template_path):
            return {"success": False, "error": f"模板文件不存在: {template_path}"}

        workbook: Workbook = load_workbook(template_path)
        sheet: Worksheet = workbook.active or workbook.worksheets[0]

        col_letter_to_idx = {}
        for field_name, col_letter in column_mapping.items():
            col_letter_to_idx[field_name] = _column_letter_to_index(col_letter)

        rows_written = 0
        for row_offset, record in enumerate(data):
            current_row = start_row + row_offset
            for field_name, col_idx in col_letter_to_idx.items():
                if field_name in record:
                    cell = sheet.cell(row=current_row, column=col_idx)
                    original_font = cell.font.copy() if cell.font else None
                    original_alignment = cell.alignment.copy() if cell.alignment else None
                    original_border = cell.border.copy() if cell.border else None
                    original_fill = cell.fill.copy() if cell.fill else None

                    cell.value = record[field_name]

                    if original_font and original_font.name:
                        cell.font = original_font
                    if original_alignment and original_alignment.horizontal:
                        cell.alignment = original_alignment
                    if original_border and original_border.style:
                        cell.border = original_border
                    if original_fill and original_fill.fgColor:
                        cell.fill = original_fill
            rows_written += 1

        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)

        workbook.save(output_path)
        workbook.close()

        return {
            "success": True,
            "rows_written": rows_written,
            "output_path": output_path,
            "file_size": os.path.getsize(output_path),
        }

    except Exception as e:
        return {"success": False, "error": str(e)}


def _column_letter_to_index(letter: str) -> int:
    """将列字母转换为列索引 (A→1, B→2, ..., Z→26, AA→27)"""
    result = 0
    for char in letter.upper():
        result = result * 26 + (ord(char) - ord("A") + 1)
    return result


def generate_generic_excel(
    template_path: str,
    data: dict[str, Any],
    output_path: str,
    cell_mapping: dict[str, str],
) -> dict[str, Any]:
    """通用 Excel 单元格填充"""
    try:
        if not os.path.exists(template_path):
            return {"success": False, "error": f"模板文件不存在: {template_path}"}

        workbook = load_workbook(template_path)
        sheet = workbook.active or workbook.worksheets[0]

        cells_updated = 0
        for data_key, cell_ref in cell_mapping.items():
            if data_key in data:
                cell = sheet[cell_ref]
                original_font = cell.font.copy() if cell.font else None
                original_alignment = cell.alignment.copy() if cell.alignment else None
                cell.value = data[data_key]
                if original_font and original_font.name:
                    cell.font = original_font
                if original_alignment and original_alignment.horizontal:
                    cell.alignment = original_alignment
                cells_updated += 1

        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)

        workbook.save(output_path)
        workbook.close()

        return {
            "success": True,
            "cells_updated": cells_updated,
            "output_path": output_path,
            "file_size": os.path.getsize(output_path),
        }

    except Exception as e:
        return {"success": False, "error": str(e)}
